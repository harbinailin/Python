import openpyxl

file_name = r'C:\Users\86159\Desktop\api.xlsx'
sheet_name = 'Sheet1'
# 打开Excel文档
book = openpyxl.load_workbook(file_name)
# 打开工作表
sh1 = book.active
# 读取工作表的属性：行数、列数
rows = sh1.max_row
cols = sh1.max_column
print("行数：{} 列数：{}".format(rows, cols))
# 读取单元格的内容
F2 = sh1['F2'].value
print('单元格F2的值：', F2)
F3 = sh1.cell(3, 6).value
print('单元格F3的值：', F3)

# 读取一行的内容 双层循环 第一层循环获取单元格 第二层循环获取单元格的内容
print('读取第一行的内容')
row1 = sh1.iter_rows(max_row=1)
for i in row1:
    for j in i:
        print(j.value, end='\t')
    print()

# 读取一列的内容
print('读取第一列的内容')
col1 = sh1.iter_cols(max_col=1)
for i in col1:
    for j in i:
        print(j.value, end='\t')
    print()
