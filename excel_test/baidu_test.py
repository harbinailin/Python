import time
import openpyxl
from selenium import webdriver

file_name = r'C:\Users\86159\Desktop\baidu.xlsx'
sh1 = 'Sheet1'
book = openpyxl.load_workbook(file_name)
sh1 = book.active
driver = webdriver.Chrome()
driver.get('https://baidu.com')

for i in range(2,sh1.max_row+1):
    time.sleep(1)
    driver.find_element_by_id('kw').send_keys(sh1.cell(i,1).value)
    time.sleep(1)
    driver.find_element_by_id('su').click()
    time.sleep(3)

    res=driver.find_element_by_xpath('//*[@id="content_left"]').text
    if sh1.cell(i,2).value in res:
        test_res='pass'
        sh1.cell(i,3).value=test_res
    else:
        test_res='false'
        sh1.cell(i,3).value=test_res
    driver.find_element_by_id('kw').clear()

book.save('baidu.xlsx')
driver.quit()
